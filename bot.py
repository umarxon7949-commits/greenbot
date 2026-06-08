"""
Telegram-бот для отчётов по таблице GREEN TASHKENT (только чтение).

Функции:
  /start, /menu      — главное меню (кнопки)
  📊 Сводка           — итоги по месяцам и категориям
  💵 Курс USD         — последний курс из листа «Курс валют»
  📦 Остаток          — остаток арматуры по диаметрам
  📞 Контакты         — поиск контакта по имени/должности
  🔄 Обновить таблицу — прислать боту новый .xlsx, бот заменит файл

Установка:
  pip install aiogram openpyxl

Запуск:
  1) Положите GREEN_TASHKENT.xlsx рядом с этим файлом (или обновите через бота).
  2) Вставьте токен из @BotFather в BOT_TOKEN.
  3) python bot.py
"""

import asyncio
import json
import os
from datetime import datetime, timedelta

import openpyxl
from aiogram import Bot, Dispatcher, F, types
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import (
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    KeyboardButton,
    ReplyKeyboardMarkup,
    WebAppInfo,
)

# ──────────────────────────── НАСТРОЙКИ ────────────────────────────
# Токен берётся из переменной окружения BOT_TOKEN (так настроено на хостинге).
# Для локального запуска можно вписать токен прямо в кавычки ниже.
BOT_TOKEN = os.getenv("BOT_TOKEN", "ВАШ_ТОКЕН_ЗДЕСЬ")

# Папка для данных. На хостинге сюда подключается постоянный диск (volume),
# чтобы обновлённый через Telegram файл не пропадал при перезапуске.
# Задаётся переменной окружения DATA_DIR (например, /data на Railway).
# Локально по умолчанию — папка рядом с bot.py.
DATA_DIR = os.getenv("DATA_DIR", os.path.dirname(__file__))
EXCEL_FILE = os.path.join(DATA_DIR, "GREEN_TASHKENT.xlsx")

# Если файла ещё нет в DATA_DIR (первый запуск на чистом volume),
# но он лежит рядом с кодом — копируем его в DATA_DIR.
_bundled = os.path.join(os.path.dirname(__file__), "GREEN_TASHKENT.xlsx")
if not os.path.exists(EXCEL_FILE) and os.path.exists(_bundled) and _bundled != EXCEL_FILE:
    import shutil
    os.makedirs(DATA_DIR, exist_ok=True)
    shutil.copy(_bundled, EXCEL_FILE)

# Ограничение доступа: впишите Telegram ID через переменную ALLOWED_USERS
# (числа через запятую, узнать ID: @userinfobot). Пусто = доступ всем.
_allowed_env = os.getenv("ALLOWED_USERS", "").strip()
ALLOWED_USERS: list[int] = (
    [int(x) for x in _allowed_env.replace(" ", "").split(",") if x]
    if _allowed_env else []
)

# Ссылка на файл в облаке (OneDrive/Google Drive и т.п.) для автообновления.
# Задаётся переменной окружения EXCEL_URL. Если пусто — автообновление выключено.
EXCEL_URL = os.getenv("EXCEL_URL", "").strip()

# Час ежедневного автообновления (по часовому поясу сервера, UTC).
# Ташкент = UTC+5, поэтому 2 UTC ≈ 7:00 утра в Ташкенте.
SYNC_HOUR_UTC = int(os.getenv("SYNC_HOUR_UTC", "2"))

# Публичный адрес мини-аппа (Railway domain). Если задан — в боте появляется
# кнопка «Открыть панель». Railway передаёт порт через переменную PORT.
WEBAPP_URL = os.getenv("WEBAPP_URL", "").strip()
PORT = int(os.getenv("PORT", "8080"))

# ──────────────────────────── БОТ ────────────────────────────
bot = Bot(token=BOT_TOKEN)
dp = Dispatcher(storage=MemoryStorage())


class Form(StatesGroup):
    contact_query = State()
    waiting_file = State()


def allowed(user_id: int) -> bool:
    return not ALLOWED_USERS or user_id in ALLOWED_USERS


def main_kb() -> ReplyKeyboardMarkup:
    rows = [
        [KeyboardButton(text="📊 Сводка"), KeyboardButton(text="💵 Курс USD")],
        [KeyboardButton(text="📦 Остаток"), KeyboardButton(text="🏗 Приход арматуры")],
        [KeyboardButton(text="🧱 Бетон"), KeyboardButton(text="📞 Контакты")],
        [KeyboardButton(text="🔄 Обновить таблицу")],
    ]
    # Если задан публичный адрес — добавляем кнопку открытия мини-аппа.
    if WEBAPP_URL:
        rows.insert(0, [KeyboardButton(
            text="📈 Открыть панель",
            web_app=WebAppInfo(url=WEBAPP_URL))])
    return ReplyKeyboardMarkup(keyboard=rows, resize_keyboard=True)


def fmt_num(value) -> str:
    """Число с пробелами-разделителями тысяч."""
    try:
        n = float(value)
    except (TypeError, ValueError):
        return str(value)
    if n == int(n):
        return f"{int(n):,}".replace(",", " ")
    return f"{n:,.2f}".replace(",", " ")


def load_wb():
    if not os.path.exists(EXCEL_FILE):
        return None
    return openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)


# ──────────────────────────── ОТЧЁТЫ ────────────────────────────
SKIP_CATS = {"Арматура тоннаж"}  # это объём (тонны), не деньги


def _num(v):
    """Число или None (отсекаем текст, ошибки вроде #DIV/0!)."""
    return float(v) if isinstance(v, (int, float)) else None


def _read_summary_data():
    """Читает лист «Итого Месяцев»: список месяцев, суммы и доллары по категориям.

    Возвращает (months, sums, usd):
      months — список названий месяцев
      sums   — {категория: [значения по месяцам в сумах]}
      usd    — {категория: [значения по месяцам в долларах]}
    """
    wb = load_wb()
    if wb is None:
        return None
    ws = wb["Итого Месяцев"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    months = [c for c in rows[1][1:] if c]
    n = len(months)

    def parse_block(start_idx):
        """Читает категории от строки start_idx до строки «Итого»."""
        data = {}
        for row in rows[start_idx:]:
            cat = row[0]
            if cat is None or str(cat).strip() == "":
                continue
            if str(cat).strip().lower() == "итого":
                break
            if cat in SKIP_CATS:
                continue
            data[cat] = [_num(v) for v in row[1:1 + n]]
        return data

    # Блок сумов начинается со строки 2 (индекс 2 = третья строка листа).
    sums = parse_block(2)
    # Блок долларов — второй такой же блок ниже. Ищем вторую шапку «Категория».
    usd = {}
    for i, row in enumerate(rows):
        if i > 1 and row and str(row[0]).strip() == "Категория":
            usd = parse_block(i + 1)
            break

    return months, sums, usd


def get_month_list():
    """Список месяцев для кнопок (только те, где есть хоть какие-то расходы)."""
    data = _read_summary_data()
    if not data:
        return []
    months, sums, _ = data
    active = []
    for idx, m in enumerate(months):
        total = 0.0
        for vals in sums.values():
            v = vals[idx] if idx < len(vals) else None
            if v is not None:
                total += v
        if total:
            active.append((idx, m))
    return active


def report_summary(month_idx=None) -> str:
    """Сводка. month_idx=None — весь период; иначе конкретный месяц."""
    data = _read_summary_data()
    if not data:
        return "⚠️ Файл таблицы не найден. Обновите его через «🔄 Обновить таблицу»."
    months, sums, usd = data

    def agg(store, idx):
        """Сумма по категории: за месяц idx или за весь период (idx=None)."""
        out = {}
        for cat, vals in store.items():
            if idx is None:
                s = sum(v for v in vals if v is not None)
            else:
                s = vals[idx] if idx < len(vals) and vals[idx] is not None else 0
            out[cat] = s
        return out

    sum_tot = agg(sums, month_idx)
    usd_tot = agg(usd, month_idx)
    grand_sum = sum(sum_tot.values())
    grand_usd = sum(usd_tot.values())

    if month_idx is None:
        header = f"📊 *СВОДКА — ВЕСЬ ПЕРИОД*\n{months[0]} — {months[-1]}"
    else:
        header = f"📊 *СВОДКА — {months[month_idx]}*"

    lines = [header, ""]
    lines.append("*Расходы по категориям:*")
    for cat, s in sorted(sum_tot.items(), key=lambda x: -x[1]):
        if s == 0:
            continue
        share = (s / grand_sum * 100) if grand_sum else 0
        d = usd_tot.get(cat, 0)
        d_str = f" / ${fmt_num(d)}" if d else ""
        lines.append(f"• {cat}: {fmt_num(s)} сум{d_str}  ({share:.1f}%)")

    if grand_sum == 0:
        lines.append("_За этот месяц расходов нет._")

    lines.append("")
    lines.append(f"*ИТОГО: {fmt_num(grand_sum)} сум*")
    if grand_usd:
        lines.append(f"*≈ ${fmt_num(grand_usd)}*")
    return "\n".join(lines)


def report_rate() -> str:
    wb = load_wb()
    if wb is None:
        return "⚠️ Файл таблицы не найден."
    ws = wb["Курс валют"]
    last = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None and row[3] is not None:
            last = row
    wb.close()
    if not last:
        return "Курс не найден."

    date_raw, cur, qty, rate = last[0], last[1], last[2], last[3]
    # Excel-дата (число) → реальная дата
    if isinstance(date_raw, (int, float)):
        date_str = (datetime(1899, 12, 30) + timedelta(days=int(date_raw))).strftime("%d.%m.%Y")
    elif isinstance(date_raw, datetime):
        date_str = date_raw.strftime("%d.%m.%Y")
    else:
        date_str = str(date_raw)

    return (f"💵 *Курс {cur}*\n\n"
            f"Дата: {date_str}\n"
            f"1 {cur} = *{fmt_num(rate)} сум*")


def report_stock() -> str:
    wb = load_wb()
    if wb is None:
        return "⚠️ Файл таблицы не найден."
    ws = wb["Остаток"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Находим строку заголовков (где первая ячейка — «Диаметр (D)»).
    head_idx = None
    for i, row in enumerate(rows):
        if row and str(row[0]).strip().startswith("Диаметр"):
            head_idx = i
            break
    if head_idx is None:
        return "Данные об остатке не найдены."

    headers = rows[head_idx]
    # Колонки: 1 = приход (итог приход), 3 = использовано (итого использован).
    # Остаток считаем как Приход − Использовано.
    # Определяем единицу: тонны, если значения дробные и небольшие.
    sample = []
    for row in rows[head_idx + 1:]:
        if row and isinstance(row[1], (int, float)):
            sample.append(row[1])
    is_tonnes = bool(sample) and all(v < 100000 for v in sample) and \
        any(v != int(v) for v in sample)
    unit = " т" if is_tonnes else " шт"

    def f(v):
        return f"{fmt_num(v)}{unit}"

    lines = ["📦 *ОСТАТОК АРМАТУРЫ*", "_(приход − использовано)_", ""]
    grand = 0.0
    any_row = False
    for row in rows[head_idx + 1:]:
        if not row or row[0] in (None, ""):
            continue
        d = str(row[0]).strip()
        if d.lower() in ("итого", "всего"):
            continue
        prihod = row[1] if isinstance(row[1], (int, float)) else 0
        ispol = row[3] if len(row) > 3 and isinstance(row[3], (int, float)) else 0
        ostatok = prihod - ispol
        grand += ostatok
        any_row = True
        lines.append(f"• Ø {d}: {f(ostatok)}")

    if not any_row:
        return "Данные об остатке не найдены."
    lines.append("")
    lines.append(f"*ИТОГО остаток: {f(grand)}*")
    return "\n".join(lines)


def _intake_by_diameter():
    """Читает лист «Арматура приход»: вес (кг) по диаметрам, сгруппированный по месяцам.

    Возвращает dict: {(год, месяц): {диаметр: тонны}} или None.
    """
    from datetime import datetime as _dt
    wb = load_wb()
    if wb is None or "Арматура приход" not in wb.sheetnames:
        return None
    ws = wb["Арматура приход"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return None

    # Находим колонки диаметров по заголовкам (Катанка, D8, D10, ... D25).
    header = rows[0]
    diam_cols = {}
    for idx, h in enumerate(header):
        if h is None:
            continue
        name = str(h).strip()
        if name == "Катанка" or (name.startswith("D") and name[1:].isdigit()):
            diam_cols[idx] = name
    if not diam_cols:
        return None

    result = {}
    for row in rows[1:]:
        d = row[1] if len(row) > 1 else None
        if not isinstance(d, _dt):
            continue
        key = (d.year, d.month)
        bucket = result.setdefault(key, {})
        for c, name in diam_cols.items():
            v = row[c] if c < len(row) else None
            if isinstance(v, (int, float)) and v > 0:
                # защита от аномалий: один прут не может весить > 50 т
                if v > 50000:
                    continue
                bucket[name] = bucket.get(name, 0) + v / 1000.0  # кг → т
    return result


_MONTH_RU = {
    1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель", 5: "Май", 6: "Июнь",
    7: "Июль", 8: "Август", 9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь",
}


def get_intake_months():
    """Месяцы, где был приход арматуры (для кнопок)."""
    data = _intake_by_diameter()
    if not data:
        return []
    out = []
    for (y, m) in sorted(data):
        if sum(data[(y, m)].values()) > 0:
            out.append((f"{y}-{m:02d}", f"{_MONTH_RU[m]} {y}"))
    return out


def report_intake(month_key=None) -> str:
    """Приход арматуры. month_key=None — весь период по месяцам;
    иначе строка 'ГГГГ-ММ' — разбивка по диаметрам за месяц."""
    data = _intake_by_diameter()
    if not data:
        return "⚠️ Данные о приходе не найдены (лист «Арматура приход»)."

    if month_key is None:
        lines = ["🏗 *ПРИХОД АРМАТУРЫ — ВЕСЬ ПЕРИОД*", ""]
        grand = 0.0
        for (y, m) in sorted(data):
            tot = sum(data[(y, m)].values())
            if tot > 0:
                grand += tot
                lines.append(f"• {_MONTH_RU[m]} {y}: {fmt_num(round(tot, 2))} т")
        lines.append("")
        lines.append(f"*ИТОГО приход: {fmt_num(round(grand, 2))} т*")
        return "\n".join(lines)

    y, m = (int(x) for x in month_key.split("-"))
    bucket = data.get((y, m), {})
    if not bucket:
        return f"🏗 *ПРИХОД — {_MONTH_RU[m]} {y}*\n\nЗа этот месяц прихода нет."

    lines = [f"🏗 *ПРИХОД АРМАТУРЫ — {_MONTH_RU[m]} {y}*", ""]
    lines.append("*По диаметрам:*")
    # сортируем по убыванию тоннажа
    for name, t in sorted(bucket.items(), key=lambda x: -x[1]):
        label = name if name == "Катанка" else f"Ø {name[1:]}"
        lines.append(f"• {label}: {fmt_num(round(t, 2))} т")
    total = sum(bucket.values())
    lines.append("")
    lines.append(f"*ИТОГО за месяц: {fmt_num(round(total, 2))} т*")
    return "\n".join(lines)


def _concrete_by_month():
    """Лист «Бетон»: по месяцам марка → [объём м³, сумма сум]."""
    from datetime import datetime as _dt
    wb = load_wb()
    if wb is None or "Бетон" not in wb.sheetnames:
        return None
    ws = wb["Бетон"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    result = {}
    for row in rows[1:]:
        d = row[0] if row else None
        if not isinstance(d, _dt):  # пропускаем строки «Итог»
            continue
        mark = row[3] if len(row) > 3 else None
        qty = row[4] if len(row) > 4 else None
        summ = row[5] if len(row) > 5 else None
        if not mark or not isinstance(qty, (int, float)):
            continue
        key = (d.year, d.month)
        bucket = result.setdefault(key, {})
        m = str(mark).strip()
        cur = bucket.setdefault(m, [0.0, 0.0])
        cur[0] += qty
        cur[1] += summ if isinstance(summ, (int, float)) else 0
    return result


def get_concrete_months():
    """Месяцы с поставками бетона (для кнопок)."""
    data = _concrete_by_month()
    if not data:
        return []
    return [(f"{y}-{m:02d}", f"{_MONTH_RU[m]} {y}") for (y, m) in sorted(data)]


def report_concrete(month_key=None) -> str:
    """Бетон: за месяц (по маркам) или весь период."""
    data = _concrete_by_month()
    if not data:
        return "⚠️ Данные по бетону не найдены (лист «Бетон»)."

    if month_key is None:
        lines = ["🧱 *БЕТОН — ВЕСЬ ПЕРИОД*", "", "*По маркам:*"]
        marks = {}
        grand_q = grand_s = 0.0
        for (y, m), bucket in data.items():
            for mark, (q, s) in bucket.items():
                cur = marks.setdefault(mark, [0.0, 0.0])
                cur[0] += q
                cur[1] += s
                grand_q += q
                grand_s += s
        for mark, (q, s) in sorted(marks.items(), key=lambda x: -x[1][0]):
            lines.append(f"• {mark}: {fmt_num(round(q, 1))} м³  ({fmt_num(round(s))} сум)")
        lines.append("")
        lines.append(f"*ИТОГО: {fmt_num(round(grand_q, 1))} м³*")
        lines.append(f"*Сумма: {fmt_num(round(grand_s))} сум*")
        return "\n".join(lines)

    y, m = (int(x) for x in month_key.split("-"))
    bucket = data.get((y, m), {})
    if not bucket:
        return f"🧱 *БЕТОН — {_MONTH_RU[m]} {y}*\n\nЗа этот месяц поставок нет."
    lines = [f"🧱 *БЕТОН — {_MONTH_RU[m]} {y}*", "", "*По маркам:*"]
    tot_q = tot_s = 0.0
    for mark, (q, s) in sorted(bucket.items(), key=lambda x: -x[1][0]):
        lines.append(f"• {mark}: {fmt_num(round(q, 1))} м³  ({fmt_num(round(s))} сум)")
        tot_q += q
        tot_s += s
    lines.append("")
    lines.append(f"*ИТОГО за месяц: {fmt_num(round(tot_q, 1))} м³*")
    lines.append(f"*Сумма: {fmt_num(round(tot_s))} сум*")
    return "\n".join(lines)


def build_app_data() -> dict:
    """Собирает все данные для мини-аппа в один словарь (как для веб-страницы)."""
    out = {"months": [], "sums": {}, "usd": {}, "intake": {}, "stock": [],
           "rate": None, "rate_date": None}

    data = _read_summary_data()
    if data:
        months, sums, usd = data
        out["months"] = months
        out["sums"] = sums
        out["usd"] = usd

    # приход по диаметрам
    intake = _intake_by_diameter()
    if intake:
        out["intake"] = {f"{y}-{m:02d}": {k: round(v, 2) for k, v in b.items()}
                         for (y, m), b in intake.items()}

    # бетон по месяцам и маркам
    concrete = _concrete_by_month()
    if concrete:
        out["concrete"] = {
            f"{y}-{m:02d}": {mark: [round(q, 1), round(s)] for mark, (q, s) in b.items()}
            for (y, m), b in concrete.items()
        }
    else:
        out["concrete"] = {}

    # остаток (приход − использовано)
    wb = load_wb()
    if wb and "Остаток" in wb.sheetnames:
        ws = wb["Остаток"]
        rows = list(ws.iter_rows(values_only=True))
        head_idx = None
        for i, row in enumerate(rows):
            if row and str(row[0]).strip().startswith("Диаметр"):
                head_idx = i
                break
        if head_idx is not None:
            for row in rows[head_idx + 1:]:
                if not row or row[0] in (None, ""):
                    continue
                d = str(row[0]).strip()
                if d.lower() in ("итого", "всего"):
                    continue
                pr = row[1] if isinstance(row[1], (int, float)) else 0
                isp = row[3] if len(row) > 3 and isinstance(row[3], (int, float)) else 0
                out["stock"].append({"d": d, "ost": pr - isp})
    if wb:
        wb.close()

    # курс
    wb = load_wb()
    if wb and "Курс валют" in wb.sheetnames:
        ws = wb["Курс валют"]
        last = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None and len(row) > 3 and row[3] is not None:
                last = row
        if last:
            out["rate"] = _num(last[3])
        wb.close()

    return out


def search_contacts(query: str) -> str:
    wb = load_wb()
    if wb is None:
        return "⚠️ Файл таблицы не найден."
    ws = wb["Контакты"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    q = query.strip().lower()
    found = []
    for row in rows:
        if not row or row[0] in (None, "№", "") or str(row[0]).startswith("📞"):
            continue
        name = str(row[1] or "")
        role = str(row[2] or "")
        phone = str(row[3] or "")
        org = str(row[4] or "")
        haystack = f"{name} {role} {org}".lower()
        if q in haystack:
            found.append(f"👤 *{name.strip()}*\n💼 {role}\n📱 {phone}\n🏢 {org}")
    if not found:
        return f"По запросу «{query}» ничего не найдено."
    return "📞 *Найдено:*\n\n" + "\n\n".join(found[:15])


# ──────────────────────────── ХЕНДЛЕРЫ ────────────────────────────
@dp.message(Command("start"))
@dp.message(Command("menu"))
async def start_handler(message: types.Message, state: FSMContext):
    await state.clear()
    if not allowed(message.from_user.id):
        await message.answer("⛔ Доступ ограничен.")
        return
    await message.answer(
        "👷 *Бот отчётов — GREEN TASHKENT*\n\nВыберите раздел:",
        reply_markup=main_kb(),
        parse_mode="Markdown",
    )


def summary_months_kb() -> InlineKeyboardMarkup:
    """Кнопки выбора месяца + «Весь период»."""
    rows = []
    months = get_month_list()
    row = []
    for idx, name in months:
        row.append(InlineKeyboardButton(text=name, callback_data=f"sum:{idx}"))
        if len(row) == 2:
            rows.append(row)
            row = []
    if row:
        rows.append(row)
    rows.append([InlineKeyboardButton(text="📈 Весь период (итого)",
                                      callback_data="sum:all")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


@dp.message(F.text == "📊 Сводка")
async def h_summary(message: types.Message):
    if not allowed(message.from_user.id):
        return
    kb = summary_months_kb()
    if not kb.inline_keyboard or len(kb.inline_keyboard) == 1:
        await message.answer(report_summary(), parse_mode="Markdown")
        return
    await message.answer("Выберите месяц или весь период:", reply_markup=kb)


@dp.callback_query(F.data.startswith("sum:"))
async def cb_summary(callback: types.CallbackQuery):
    if not allowed(callback.from_user.id):
        await callback.answer()
        return
    key = callback.data.split(":", 1)[1]
    idx = None if key == "all" else int(key)
    await callback.message.answer(report_summary(idx), parse_mode="Markdown")
    await callback.answer()


@dp.message(F.text == "💵 Курс USD")
async def h_rate(message: types.Message):
    if not allowed(message.from_user.id):
        return
    await message.answer(report_rate(), parse_mode="Markdown")


@dp.message(F.text == "📦 Остаток")
async def h_stock(message: types.Message):
    if not allowed(message.from_user.id):
        return
    await message.answer(report_stock(), parse_mode="Markdown")


def intake_months_kb() -> InlineKeyboardMarkup:
    rows = []
    row = []
    for idx, name in get_intake_months():
        row.append(InlineKeyboardButton(text=name, callback_data=f"in:{idx}"))
        if len(row) == 2:
            rows.append(row)
            row = []
    if row:
        rows.append(row)
    rows.append([InlineKeyboardButton(text="📈 Весь период (итого)",
                                      callback_data="in:all")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


@dp.message(F.text == "🏗 Приход арматуры")
async def h_intake(message: types.Message):
    if not allowed(message.from_user.id):
        return
    kb = intake_months_kb()
    if len(kb.inline_keyboard) == 1:  # только кнопка «весь период»
        await message.answer(report_intake(), parse_mode="Markdown")
        return
    await message.answer("Выберите месяц или весь период:", reply_markup=kb)


@dp.callback_query(F.data.startswith("in:"))
async def cb_intake(callback: types.CallbackQuery):
    if not allowed(callback.from_user.id):
        await callback.answer()
        return
    key = callback.data.split(":", 1)[1]
    month_key = None if key == "all" else key
    await callback.message.answer(report_intake(month_key), parse_mode="Markdown")
    await callback.answer()


def concrete_months_kb() -> InlineKeyboardMarkup:
    rows = []
    row = []
    for key, name in get_concrete_months():
        row.append(InlineKeyboardButton(text=name, callback_data=f"bet:{key}"))
        if len(row) == 2:
            rows.append(row)
            row = []
    if row:
        rows.append(row)
    rows.append([InlineKeyboardButton(text="📈 Весь период (итого)",
                                      callback_data="bet:all")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


@dp.message(F.text == "🧱 Бетон")
async def h_concrete(message: types.Message):
    if not allowed(message.from_user.id):
        return
    kb = concrete_months_kb()
    if len(kb.inline_keyboard) == 1:
        await message.answer(report_concrete(), parse_mode="Markdown")
        return
    await message.answer("Выберите месяц или весь период:", reply_markup=kb)


@dp.callback_query(F.data.startswith("bet:"))
async def cb_concrete(callback: types.CallbackQuery):
    if not allowed(callback.from_user.id):
        await callback.answer()
        return
    key = callback.data.split(":", 1)[1]
    month_key = None if key == "all" else key
    await callback.message.answer(report_concrete(month_key), parse_mode="Markdown")
    await callback.answer()


@dp.message(F.text == "📞 Контакты")
async def h_contacts(message: types.Message, state: FSMContext):
    if not allowed(message.from_user.id):
        return
    await state.set_state(Form.contact_query)
    await message.answer("Введите имя, должность или организацию для поиска:")


@dp.message(Form.contact_query)
async def h_contacts_query(message: types.Message, state: FSMContext):
    await state.clear()
    await message.answer(search_contacts(message.text), parse_mode="Markdown",
                         reply_markup=main_kb())


@dp.message(F.text == "🔄 Обновить таблицу")
async def h_update(message: types.Message, state: FSMContext):
    if not allowed(message.from_user.id):
        return
    await state.set_state(Form.waiting_file)
    await message.answer("Пришлите новый файл *GREEN_TASHKENT.xlsx* как документ.",
                         parse_mode="Markdown")


@dp.message(Form.waiting_file, F.document)
async def h_receive_file(message: types.Message, state: FSMContext):
    await state.clear()
    doc = message.document
    if not doc.file_name.lower().endswith(".xlsx"):
        await message.answer("❌ Это не .xlsx файл.", reply_markup=main_kb())
        return
    file = await bot.get_file(doc.file_id)
    await bot.download_file(file.file_path, EXCEL_FILE)
    await message.answer("✅ Таблица обновлена.", reply_markup=main_kb())


@dp.message(Form.waiting_file)
async def h_receive_file_wrong(message: types.Message):
    await message.answer("Жду файл .xlsx как документ. Или /menu для отмены.")


@dp.message()
async def fallback(message: types.Message):
    if not allowed(message.from_user.id):
        return
    await message.answer("Выберите раздел в меню 👇", reply_markup=main_kb())


# ──────────────────────────── АВТООБНОВЛЕНИЕ ИЗ ОБЛАКА ────────────────────────────
import urllib.request
import base64


def _to_direct_url(url: str) -> str:
    """Превращает ссылку OneDrive/Google Drive в прямую загрузку файла."""
    u = url.strip()
    # OneDrive личный (1drv.ms / onedrive.live.com) → base64-приём для прямой загрузки
    if "1drv.ms" in u or "onedrive.live.com" in u:
        b64 = base64.urlsafe_b64encode(u.encode()).decode().rstrip("=")
        return f"https://api.onedrive.com/v1.0/shares/u!{b64}/root/content"
    # SharePoint / OneDrive for Business: добавляем download=1
    if "sharepoint.com" in u or "-my.sharepoint" in u:
        sep = "&" if "?" in u else "?"
        return u + sep + "download=1"
    # Google Drive: /file/d/<ID>/view → uc?export=download&id=<ID>
    if "drive.google.com" in u and "/d/" in u:
        try:
            fid = u.split("/d/")[1].split("/")[0]
            return f"https://drive.google.com/uc?export=download&id={fid}"
        except IndexError:
            return u
    return u


def sync_from_url() -> tuple[bool, str]:
    """Скачивает файл по EXCEL_URL и сохраняет в EXCEL_FILE. (успех, сообщение)."""
    if not EXCEL_URL:
        return False, "Ссылка для автообновления не задана (переменная EXCEL_URL)."
    direct = _to_direct_url(EXCEL_URL)
    try:
        req = urllib.request.Request(direct, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=60) as resp:
            data = resp.read()
        # Проверяем, что это xlsx (zip-архив начинается с 'PK').
        if data[:2] != b"PK":
            return False, ("Скачан не Excel-файл. Проверьте, что ссылка ведёт "
                           "на сам файл и открыта для всех по ссылке.")
        os.makedirs(DATA_DIR, exist_ok=True)
        with open(EXCEL_FILE, "wb") as f:
            f.write(data)
        return True, f"Таблица обновлена из облака ({len(data) // 1024} КБ)."
    except Exception as e:
        return False, f"Не удалось скачать: {e}"


async def daily_sync_loop():
    """Раз в сутки в SYNC_HOUR_UTC скачивает свежий файл из облака."""
    if not EXCEL_URL:
        return
    while True:
        now = datetime.utcnow()
        target = now.replace(hour=SYNC_HOUR_UTC, minute=0, second=0, microsecond=0)
        if target <= now:
            target += timedelta(days=1)
        await asyncio.sleep((target - now).total_seconds())
        ok, msg = sync_from_url()
        print(f"[автообновление] {msg}")


@dp.message(Command("sync"))
async def h_sync(message: types.Message):
    """Ручной запуск обновления из облака — проверить, что ссылка работает."""
    if not allowed(message.from_user.id):
        return
    await message.answer("⏳ Скачиваю файл из облака…")
    ok, msg = sync_from_url()
    await message.answer(("✅ " if ok else "❌ ") + msg, reply_markup=main_kb())


# ──────────────────────────── ВЕБ-СЕРВЕР (МИНИ-АПП) ────────────────────────────
from aiohttp import web

# Папка, где лежит miniapp.html (рядом с bot.py).
_APP_DIR = os.path.dirname(__file__)


async def web_index(request):
    # Файл мини-аппа может называться index.html или miniapp.html.
    for fname in ("index.html", "miniapp.html"):
        path = os.path.join(_APP_DIR, fname)
        if os.path.exists(path):
            return web.FileResponse(path)
    return web.Response(text="Файл мини-аппа не найден", status=404)


async def web_data(request):
    """Отдаёт данные таблицы в JSON для мини-аппа."""
    try:
        data = build_app_data()
    except Exception as e:
        return web.json_response({"error": str(e)}, status=500)
    return web.json_response(data, headers={"Cache-Control": "no-store"})


async def web_health(request):
    return web.Response(text="ok")


def make_web_app() -> web.Application:
    app = web.Application()
    app.router.add_get("/", web_index)
    app.router.add_get("/api/data", web_data)
    app.router.add_get("/health", web_health)
    return app


# ──────────────────────────── ЗАПУСК ────────────────────────────
async def main():
    print("Бот запущен.")
    if EXCEL_URL:
        asyncio.create_task(daily_sync_loop())
        print(f"Автообновление включено: каждый день в {SYNC_HOUR_UTC}:00 UTC.")

    # Поднимаем веб-сервер (для мини-аппа) параллельно с ботом.
    runner = web.AppRunner(make_web_app())
    await runner.setup()
    site = web.TCPSite(runner, "0.0.0.0", PORT)
    await site.start()
    print(f"Веб-сервер мини-аппа слушает порт {PORT}.")

    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())
